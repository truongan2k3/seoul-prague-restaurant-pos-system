#!/usr/bin/env python3
"""Generate supabase/seed-menu-from-old-system.sql from menu-old-system.xlsx.

Usage:
  python3 scripts/generate-menu-from-xlsx.py [/path/to/menu-old-system.xlsx]

Requires: pip install openpyxl
"""

from __future__ import annotations

import hashlib
import re
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError as exc:
    raise SystemExit("Install openpyxl: pip install openpyxl") from exc

ROOT = Path(__file__).resolve().parents[1]
DEFAULT_XLSX = Path.home() / "Downloads" / "menu-old-system.xlsx"
OUT_SQL = ROOT / "supabase" / "seed-menu-from-old-system.sql"
LUNCH_SQL = ROOT / "supabase" / "seed-lunch-menu.sql"
JIN_Cheng_SCRIPT = ROOT / "scripts" / "generate-menu-seed.mjs"

DRINK_CATEGORIES = {
    "Domácí limonáda & ledový čaj",
    "Na čepu/Nealkohol",
    "Kafe & čaj",
    "Koktejly",
    "Korejské nápoje",
    "Shots",
    "Vino",
}

# Chinese fallbacks for drinks / items missing zh in spreadsheet.
ZH_FALLBACK: dict[str, str] = {
    "malinovka": "覆盆子",
    "yuzu & citron": "柚子与柠檬",
    "mango": "芒果",
    "jasmín & citron ledový čaj": "茉莉柠檬冰茶",
    "červený čaj s liči": "荔枝红茶冰茶",
    "kofola 0,5l": "Kofola 0.5L",
    "kofola 0,3l": "Kofola 0.3L",
    "pilsner urquell 12° 0,3l": "Pilsner Urquell 0.3L",
    "pilsner urquell 12° 0,5l": "Pilsner Urquell 0.5L",
    "coca cola / cola zero / fanta / sprite": "可口可乐/零度/芬达/雪碧",
    "romequell 0,75l": "Romequell 0.75L",
    "voda": "水",
    "karafa vody": "一壶水",
    "espresso / americano / lungo": "浓缩/美式/长萃",
    "latte / cappucino": "拿铁/卡布奇诺",
    "jasmín": "茉莉花茶",
    "oolong": "乌龙茶",
    "červený čaj": "红茶",
    "yuzu highball": "柚子嗨棒",
    "soju mojito": "烧酒莫吉托",
    "sakura gin fizz": "櫻花金酒菲士",
    "seoul sunset": "首尔日落",
    "makgeolli peach punch": "米酒蜜桃潘趣",
    "soju 0,33l": "韩国烧酒 0.33L",
    "makgeolli 0,75l": "马格利米酒 0.75L",
    "milkis": "Milkis 乳酸菌饮料",
    "finlandia (shot)": "芬兰伏特加",
    "jägermeister (shot)": "野格",
    "gin (shot)": "金酒",
    "beefeater (shot)": "必富达金酒",
    "jameson (shot)": "詹姆森威士忌",
    "jim beam (shot)": "占边威士忌",
    "rum (shot)": "朗姆酒",
    "martini bianco (shot)": "马天尼 bianco",
    "chardonnay (2dl)": "霞多丽葡萄酒",
    "cabernet sauvignon (2dl)": "赤霞珠葡萄酒",
    "box": "BOX",
}


def normalize_key(value: str) -> str:
    return re.sub(r"\s+", " ", value.lower().replace("·", "-").strip())


def sql_str(value: str | None) -> str:
    if value is None:
        return "NULL"
    text = str(value).strip()
    if not text:
        return "NULL"
    return "'" + text.replace("'", "''") + "'"


def stable_uuid(category: str, name_cz: str, name_zh: str, sort_order: int) -> str:
    digest = hashlib.sha1(
        f"{category}\0{name_cz}\0{name_zh}\0{sort_order}".encode("utf-8")
    ).hexdigest()
    return f"50000000-{digest[:4]}-4000-a000-{digest[4:16]}"


def parse_lunch_seed() -> dict[str, tuple[str, str, str]]:
    text = LUNCH_SQL.read_text(encoding="utf-8")
    pattern = re.compile(r"\('([^']*)',\s*'([^']*)',\s*'([^']*)'")
    mapping: dict[str, tuple[str, str, str]] = {}
    for match in pattern.finditer(text):
        en, cz, zh = match.group(1), match.group(2), match.group(3)
        mapping[normalize_key(cz)] = (en, cz, zh)
    return mapping


def parse_jin_cheng_csv() -> dict[str, tuple[str, str, str]]:
    text = JIN_Cheng_SCRIPT.read_text(encoding="utf-8")
    block = text.split('const rows = `', 1)[1].split('`;', 1)[0]
    mapping: dict[str, tuple[str, str, str]] = {}
    for line in block.strip().split("\n"):
        parts = []
        current = ""
        in_quotes = False
        for ch in line:
            if ch == '"':
                in_quotes = not in_quotes
            elif ch == "," and not in_quotes:
                parts.append(current)
                current = ""
            else:
                current += ch
        parts.append(current)
        if len(parts) < 5:
            continue
        _category, _price, name_en, name_cz, name_zh = parts[:5]
        mapping[normalize_key(name_cz)] = (name_en, name_cz, name_zh)
        mapping[normalize_key(name_en)] = (name_en, name_cz, name_zh)
    return mapping


def resolve_names(
    name_cz: str,
    name_zh: str,
    lunch_map: dict[str, tuple[str, str, str]],
    jin_map: dict[str, tuple[str, str, str]],
) -> tuple[str, str, str]:
    key = normalize_key(name_cz)
    if not name_zh:
        name_zh = ZH_FALLBACK.get(key, "")

    for source in (lunch_map, jin_map):
        if key in source:
            en, cz, zh = source[key]
            return en, cz, zh or name_zh

    # Partial match (lunch uses "Kung Pao · kuřecí", excel "Kung Pao - Kuřecí")
    for source in (lunch_map, jin_map):
        for lk, (en, cz, zh) in source.items():
            if lk in key or key in lk:
                return en, cz, name_zh or zh

    # Derive English from Czech label patterns
    en = name_cz
    replacements = [
        (r" - Kuřecí$", " Chicken"),
        (r" - Vepřové$", " Pork"),
        (r" - Hovězí$", " Beef"),
        (r" - Krevety$", " Prawns"),
        (r" - Tofu$", " Tofu"),
        (r"^Kung Pao", "Kung Pao"),
        (r"^Nudle", "Noodles"),
    ]
    if " - " in name_cz:
        base, protein = name_cz.split(" - ", 1)
        protein_en = {
            "kuřecí": "Chicken",
            "vepřové": "Pork",
            "hovězí": "Beef",
            "krevety": "Prawns",
            "tofu": "Tofu",
        }.get(protein.lower(), protein)
        en = f"{base} ({protein_en})"

    return en, name_cz, name_zh or name_cz


def read_xlsx(path: Path) -> list[dict]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    items: list[dict] = []
    for row in range(2, ws.max_row + 1):
        category = str(ws.cell(row, 1).value or "").strip()
        name_zh = str(ws.cell(row, 2).value or "").strip()
        name_cz = str(ws.cell(row, 3).value or "").strip()
        price = ws.cell(row, 4).value
        if not category or (not name_cz and not name_zh):
            continue
        if not name_cz:
            name_cz = name_zh
        items.append(
            {
                "category": category,
                "name_zh": name_zh,
                "name_cz": name_cz,
                "price": float(price or 0),
            }
        )
    return items


def category_type(category: str) -> str:
    return "drink" if category in DRINK_CATEGORIES else "dish"


def category_station(category: str) -> str:
    return "bar" if category in DRINK_CATEGORIES else "kitchen"


def tax_group(category: str) -> str:
    return "A" if category in DRINK_CATEGORIES else "B"


def generate_sql(items: list[dict]) -> str:
    lunch_map = parse_lunch_seed()
    jin_map = parse_jin_cheng_csv()

    categories: list[str] = []
    for item in items:
        if item["category"] not in categories:
            categories.append(item["category"])

    cat_value_rows: list[str] = []
    for index, category in enumerate(categories):
        cat_type = category_type(category)
        cat_value_rows.append(
            f"  ({sql_str(category)}, {sql_str(cat_type)}, {index + 1})"
        )

    cat_insert = ""
    if cat_value_rows:
        cat_insert = "INSERT INTO public.categories (name, type, display_order) VALUES\n"
        cat_insert += ",\n".join(cat_value_rows)
        cat_insert += ";"

    value_rows: list[str] = []
    for sort_order, raw in enumerate(items, start=1):
        name_en, name_cz, name_zh = resolve_names(
            raw["name_cz"], raw["name_zh"], lunch_map, jin_map
        )
        category = raw["category"]
        item_id = stable_uuid(category, name_cz, name_zh, sort_order)
        station = category_station(category)
        item_type = "drink" if station == "bar" else "food"
        tax = tax_group(category)

        value_rows.append(
            f"""  (
    '{item_id}'::uuid,
    {sql_str(category)},
    {raw["price"]},
    {sql_str(name_en)},
    {sql_str(name_cz)},
    {sql_str(name_zh)},
    NULL,
    NULL,
    NULL,
    NULL,
    true,
    {sql_str(station)},
    {sql_str(item_type)},
    {sort_order},
    {sort_order},
    {sql_str(name_en)},
    NULL,
    false,
    {sql_str(tax)}
  )"""
        )

    values_sql = ",\n".join(value_rows)

    return f"""-- Full menu from menu-old-system.xlsx ({len(items)} items, {len(categories)} categories)
-- Regenerate: python3 scripts/generate-menu-from-xlsx.py
-- Run in Supabase SQL Editor AFTER patch-menu-jin-cheng-schema.sql and patch-tax-summary.sql
-- WARNING: Deletes ALL menu_items and categories, then loads Excel menu only.
-- Do NOT run seed-lunch-menu.sql or seed-menu-jin-cheng.sql afterward.

BEGIN;

DELETE FROM public.menu_items;
DELETE FROM public.categories;

{cat_insert}

INSERT INTO public.menu_items (
  id,
  category,
  price,
  name_en,
  name_cz,
  name_zh,
  description_en,
  description_cz,
  description_zh,
  image_url,
  is_available,
  station,
  item_type,
  sort_order,
  display_order,
  name,
  description,
  sold_out,
  tax_group
) VALUES
{values_sql};

UPDATE public.menu_items mi
SET category_id = c.id
FROM public.categories c
WHERE lower(trim(mi.category)) = lower(trim(c.name));

COMMIT;
"""


def update_lunch_chinese(items: list[dict]) -> None:
    if not LUNCH_SQL.exists():
        return
    text = LUNCH_SQL.read_text(encoding="utf-8")
    excel_by_cz: dict[str, str] = {}
    for item in items:
        if item["name_zh"]:
            excel_by_cz[normalize_key(item["name_cz"])] = item["name_zh"]

    def repl(match: re.Match[str]) -> str:
        en, cz, zh = match.group(1), match.group(2), match.group(3)
        key = normalize_key(cz)
        new_zh = excel_by_cz.get(key)
        if not new_zh:
            for ek, ev in excel_by_cz.items():
                if ek in key or key in ek:
                    new_zh = ev
                    break
        if not new_zh or new_zh == zh:
            return match.group(0)
        return f"('{en}', '{cz}', '{new_zh}'"

    updated = re.sub(
        r"\('([^']*)',\s*'([^']*)',\s*'([^']*)'",
        repl,
        text,
    )
    if updated != text:
        LUNCH_SQL.write_text(updated, encoding="utf-8")
        print(f"Updated Chinese names in {LUNCH_SQL}")


def main() -> None:
    xlsx_path = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_XLSX
    if not xlsx_path.exists():
        raise SystemExit(f"XLSX not found: {xlsx_path}")

    items = read_xlsx(xlsx_path)
    sql = generate_sql(items)
    OUT_SQL.write_text(sql, encoding="utf-8")
    update_lunch_chinese(items)
    missing_zh = sum(1 for i in items if not resolve_names(i["name_cz"], i["name_zh"], parse_lunch_seed(), parse_jin_cheng_csv())[2])
    print(f"Wrote {len(items)} items to {OUT_SQL}")
    print(f"Items still missing Chinese after fallbacks: {missing_zh}")


if __name__ == "__main__":
    main()
